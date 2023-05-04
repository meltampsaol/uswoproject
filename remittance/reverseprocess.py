import mysql.connector
import pandas as pd
from decimal import Decimal, getcontext
import openpyxl

# Set the decimal precision to 2
ctx = getcontext()
ctx.prec = 2

# Connect to the MySQL database
cnx = mysql.connector.connect(user='root', password='mbt073233',
                              host='localhost', database='uswo_project2923')
cursor = cnx.cursor()

# Define the query to retrieve the data from the f4detail table
query = "SELECT * FROM f4detail WHERE reported = '06-2023' AND did = 9"
cursor.execute(query)
result = cursor.fetchall()

# Convert the result set to a pandas DataFrame
df_f4 = pd.DataFrame(result, columns=cursor.column_names)

# Load the Excel file and select the worksheet
filename = r'MICOGN_template.xlsx'
book = openpyxl.load_workbook(filename)
sheet = book['OGNF4Details']

lcode_list = list(df_f4['lcode'])
for i, row in df_f4.iterrows():
    for j, col in enumerate(df_f4.columns):
        if col not in ['totaloffering', 'offrefund', 'exptotal', 'remainder', 'rtotal']:
            sheet.cell(row=i+2, column=j+1, value=row[col])
            sheet.cell(row=i+2, column=j+1).number_format = '0.00'

# Save the updated Excel file
book.save(filename)

# Close the database connection
cursor.close()
cnx.close()
