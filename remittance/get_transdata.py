import openpyxl
import mysql.connector
import sys
# Establishing the connection
conn = mysql.connector.connect(
    user='root', password='mbt073233', host='127.0.0.1', database='uswo_project2923')


lcode = sys.argv[1]
xlcode = int(lcode)

# convert the xlcode string to an integer
# xlcode = int(lcodestr)
# Load Excel data
filename = "MICOGN_week17_2023.xlsx"
wb = openpyxl.load_workbook(filename=filename, read_only=True, data_only=True)
ws = wb["MICF4Details"]

# Find the row number that satisfies the condition lcode=args.lcode
row_number = None
for row in ws.iter_rows(min_row=5, min_col=28, max_col=28):
    if row[0].value == xlcode:
        row_number = row[0].row
        break

# Get the values of uswo and cfototal
if row_number is not None:
    uswo = ws["T" + str(row_number)].value or 0.00
    cfototal = ws["F" + str(row_number)].value * \
    0.15 if ws["F" + str(row_number)].value else 0.00
    cfolocale = ws["F" + str(row_number)].value * \
    0.10 if ws["F" + str(row_number)].value else 0.00
    cfointl = ws["F" + str(row_number)].value * \
    0.05 if ws["F" + str(row_number)].value else 0.00
    lcode = ws["AB" + str(row_number)].value
    did = ws["AC" + str(row_number)].value  
    wkno = ws["AE" + str(row_number)].value


cursor = conn.cursor()
 # Preparing SQL query to INSERT a record into the database.
sql = """INSERT INTO remit (
                lcode, did, wkno, cfremit, lfremit, cfintl
            ) VALUES (
                %s, %s, %s, %s, %s, %s
            )"""
values = (lcode, did, wkno, cfototal, uswo, cfointl)
try:
        # Executing the SQL command
        result = cursor.execute(sql, values)
        print("Record inserted successfully")
        conn.commit()
except Exception as e:
        
        conn.rollback()

# Closing the connection
conn.close()
wb.close()
